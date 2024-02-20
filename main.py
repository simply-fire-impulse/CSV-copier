from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import gspread
import time

st = time.time()
gc = gspread.service_account()
new_spreadsheet = gc.create(title='trial_1')
new_worksheet = new_spreadsheet.add_worksheet(title='xyz', rows=100, cols=100)
ws1 = new_spreadsheet.worksheets()[0]
new_spreadsheet.del_worksheet(ws1)

def copy_sheet_with_format(input_files, output_file):
    # Create a new workbook and worksheet
    output_wb = Workbook()
    
    # Combine worksheet names from all input files
    ws_list = []
    for input_file in input_files:
        input_wb = load_workbook(input_file)
        ws_list.extend(input_wb.sheetnames)

    # Create worksheets in the output workbook
    for i, ws_name in enumerate(ws_list, start=1):
        output_wb.create_sheet(title=f"Sheet_page{i}")
    output_wb.remove(output_wb['Sheet'])

    o = 0
    for input_file in input_files:
        input_wb = load_workbook(input_file)
        for ws_name in input_wb.sheetnames:
            ws = input_wb[ws_name]
            for row in ws.iter_rows(values_only=True):
                output_wb.worksheets[o].append(row)
            # Copy merged cells
            for merged_cell_range in ws.merged_cells.ranges:
                output_wb.worksheets[o].merge_cells(str(merged_cell_range.coord))
            o += 1

    for ws in output_wb.worksheets:
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        new_ws = new_spreadsheet.add_worksheet(title=ws.title, rows=len(data), cols=len(data[0]))
        new_ws.update(data, 'A1')

        # Merge cells in the new worksheet based on extracted merged cell information
        for merged_range in ws.merged_cells.ranges:
            start_row, start_col, end_row, end_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            new_ws.merge_cells(start_row, start_col, end_row, end_col)
    
    # Save the output workbook
    output_wb.save(output_file)
    print("Sheet copied successfully with format and merged cells.")

# Example usage:
input_files = ["./resources/ws1.xlsx", "./resources/ws2.xlsx"]
output_file = "./resources/output.xlsx"

start_time = time.time()
copy_sheet_with_format(input_files, output_file)
end_time = time.time()
print("Time taken:", end_time - start_time, "seconds")

new_spreadsheet.del_worksheet(new_worksheet)
new_spreadsheet.share('ashmit.gupta@impulsecompute.com', perm_type='user', role='writer')
et = time.time()

print(f"{et-st} s")