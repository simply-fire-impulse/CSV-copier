from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

def copy_sheet_with_format(input_file, input_sheet_name, output_file, output_sheet_name):
    # Load the input workbook
    input_wb = load_workbook(input_file)
    input_ws = input_wb[input_sheet_name]
    
    # Create a new workbook and worksheet
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = output_sheet_name
    
    # Copy values and formats
    for row in input_ws.iter_rows(values_only=True):
        output_ws.append(row)
    
    # Copy cell styles
    for row in input_ws.iter_rows(min_row=1, max_row=input_ws.max_row, min_col=1, max_col=input_ws.max_column):
        for cell in row:
            output_ws[cell.coordinate].style = cell.style
    
    # Copy merged cells
    for merged_cell_range in input_ws.merged_cells.ranges:
        print(merged_cell_range.coord)
        output_ws.merge_cells(str(merged_cell_range.coord))
    
    output_ws.font = Font(size=10)
    # Save the output workbook
    output_wb.save(output_file)
    print("Sheet copied successfully with format and merged cells.")

# Example usage:
input_file = "./resources/ws1.xlsx"
input_sheet_name = "Sheet1"
output_file = "./resources/ws2.xlsx"
output_sheet_name = "Sheet1"

copy_sheet_with_format(input_file, input_sheet_name, output_file, output_sheet_name)

